"""
Excel spreadsheet generator for the call index.

One row per call:
  #, Date/Time, Inmate, Outside Number, Facility, Filename, Duration,
  Outcome, Notes, Summary, Full Transcript
"""

import logging
import os
from typing import List, Optional
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

HEADERS = [
    "#", "Date/Time", "Inmate", "Outside Number", "Facility",
    "Filename", "Duration", "Outcome", "Notes", "Summary", "Full Transcript",
]

HEADER_FILL = PatternFill("solid", fgColor="1B2D4A")  # dark navy
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

ALT_FILL = PatternFill("solid", fgColor="FAF8F3")  # warm cream
ERR_ALT_FILL = PatternFill("solid", fgColor="FEF2F2")  # red-50

COL_WIDTHS = {
    "A": 6,    # #
    "B": 18,   # Date/Time
    "C": 20,   # Inmate
    "D": 16,   # Outside Number
    "E": 20,   # Facility
    "F": 35,   # Filename
    "G": 10,   # Duration
    "H": 18,   # Outcome
    "I": 30,   # Notes
    "J": 80,   # Summary
    "K": 40,   # Full Transcript
}

THIN = Side(border_style="thin", color="E0DACE")
CELL_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def _format_duration(seconds: Optional[float]) -> str:
    if not seconds:
        return ""
    secs = int(seconds)
    h, rem = divmod(secs, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h}:{m:02d}:{s:02d}"
    return f"{m}:{s:02d}"


def _transcript_text(turns) -> str:
    if not turns:
        return ""
    return "\n".join(f"{t.speaker}: {t.text}" for t in turns)



def generate_excel(calls, error_calls=None) -> bytes:
    """
    Generate Excel workbook from a list of CallResult objects.
    Optionally includes an Errors sheet for failed calls.
    Returns bytes of the .xlsx file.
    """
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Call Index"

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = CELL_BORDER
    ws.row_dimensions[1].height = 22

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    # Data rows
    for row_idx, call in enumerate(calls, start=2):
        is_alt = row_idx % 2 == 0
        fill = ALT_FILL if is_alt else PatternFill()

        def write(col, value):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = CELL_BORDER
            if is_alt:
                cell.fill = fill
            return cell

        write(1, call.index + 1)
        write(2, call.call_datetime_str or "")
        write(3, call.inmate_name or "")
        write(4, call.outside_number_fmt or "")
        write(5, call.facility or "")
        
        # Link filename to viewer
        fn_cell = write(6, call.filename)
        if call.mp3_path:
            audio_name = os.path.basename(call.mp3_path)
        else:
            # Fallback: swap .wav extension to .mp3 since viewer only has MP3s
            base = call.filename
            if base.lower().endswith(".wav"):
                base = base[:-4] + ".mp3"
            audio_name = base
        fn_cell.hyperlink = f"viewer.html?call={quote(audio_name)}"
        fn_cell.font = Font(underline="single", color="1B2D4A")

        write(7, _format_duration(call.duration_seconds))
        write(8, call.call_outcome or "")
        write(9, call.notes or "")
        write(10, call.summary or "")
        write(11, _transcript_text(call.turns))

        # Row height based on summary length
        summary_len = len(call.summary or "")
        height = max(30, min(120, 15 + summary_len // 4))
        ws.row_dimensions[row_idx].height = height

    # Column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Errors sheet ──
    if error_calls:
        err_ws = wb.create_sheet("Errors")
        err_headers = ["#", "Filename", "Date/Time", "Inmate", "Stage", "Error"]
        for col_idx, header in enumerate(err_headers, start=1):
            cell = err_ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = PatternFill("solid", fgColor="7F1D1D")  # red-900
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = CELL_BORDER
        err_ws.row_dimensions[1].height = 22
        err_ws.freeze_panes = "A2"

        for row_idx, call in enumerate(error_calls, start=2):
            def err_write(col, value):
                cell = err_ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = CELL_BORDER
                if row_idx % 2 == 0:
                    cell.fill = ERR_ALT_FILL
                return cell

            err_write(1, call.index + 1)
            err_write(2, call.filename)
            err_write(3, call.call_datetime_str or "")
            err_write(4, call.inmate_name or "")
            err_write(5, call.status.value if hasattr(call.status, 'value') else str(call.status))
            err_write(6, call.error or "Unknown error")
            err_ws.row_dimensions[row_idx].height = 30

        err_col_widths = {"A": 6, "B": 35, "C": 18, "D": 20, "E": 16, "F": 60}
        for col_letter, width in err_col_widths.items():
            err_ws.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
